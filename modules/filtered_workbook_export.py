from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modules.google_sheets_sync import DISPLAY_EXCLUDED_COLUMNS, _fetch_financials_sheet_rows, nice_sheet_header
from modules.utils import flatten_for_sheet, format_industry_codes

TITLE_FILL = PatternFill("solid", fgColor="1C5C5C")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
THIN_SIDE = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

EXPORT_EXCLUDED_COLUMNS = set(DISPLAY_EXCLUDED_COLUMNS) | {
    "phone",
    "lei",
    "recommended_action",

    # Old OpenRegister temporary/legacy naming.
    "openregister_assets_eur",

    # Legacy shareholder/owner integration fields.
    "relation_start_date",
    "relation_start_year",
    "main_owner_year_integrated",

    # Removed from new export format.
    "shareholder_contribution",
    "has_sole_owner",
    "has_representative_owner",
    "owner_managed",
    "is_family_owned",
    "has_majority_owner",
    "largest_owner_percentage",
    "main_owner_name",
    "main_owner_type",
    "main_owner_percentage_share",
    "main_ubo_name",
    "main_ubo_age",
    "main_ubo_percentage_share",
    "main_ubo_max_percentage_share",

    # Legacy mixed business/financial fields.
    # These should never appear as NorthData columns.
    "purpose",
    "employees",
    "balance_sheet_total_eur",
    "net_income_eur",
    "equity_eur",
    "cash_eur",
    "liabilities_eur",
    "real_estate_eur",
    "capital_amount_eur",
    "financials_date",

    # Legacy owner aggregate names.
    "number_of_owners",
    "natural_person_owner_count",
    "legal_person_owner_count",
    "youngest_owner_age",
    "oldest_owner_age",
}

PREFERRED_COLUMN_ORDER = {
    "Overview": [
        "openregister_company_id",
        "company_name",
        "legal_form",
        "founding_year",
        "active",
        "country",
        "register_number",
        "register_court",
        "register_type",
        "city",
        "postal_code",
        "website",
        "email",

        "northdata_wz_code",
        "openregister_wz_codes",
        "divisions",

        "claude_business_segment",
        "claude_assumption",

        "northdata_business_model",
        "claude_business_model",
        "claude_detailed_business_summary",

        "northdata_revenue_eur",
        "openregister_revenue_eur",

        "northdata_employees",
        "openregister_employees",

        "northdata_balance_sheet_total_eur",
        "openregister_balance_sheet_total_eur",

        "northdata_net_income_eur",
        "openregister_net_income_eur",

        "northdata_equity_eur",
        "openregister_equity_eur",

        "northdata_cash_eur",
        "openregister_cash_eur",

        "northdata_liabilities_eur",
        "openregister_liabilities_eur",

        "northdata_financials_date",
        "openregister_financials_date",

        "shareholder_name",
        "shareholder_age",
        "shareholder_type",
        "shareholder_ownership_percentage",

        "number_of_shareholders",
        "natural_person_shareholder_count",
        "legal_person_shareholder_count",
        "youngest_shareholder_age",
        "oldest_shareholder_age",

        "ubo_name",
        "ubo_age",
        "ubo_type",
        "ubo_percentage_share",
        "ubo_max_percentage_share",
        "ubo_count",
        "oldest_ubo_age",
        "youngest_ubo_age",

        "fit_score",
        "fit_label",
        "fit_comment",
    ],

    "Companies": [
        "openregister_company_id",
        "name",
        "legal_form",
        "founding_year",
        "active",
        "status",
        "country",
        "register_number",
        "register_court",
        "register_type",
        "city",
        "postal_code",
        "street",
        "formatted_address",
        "website",
        "email",
        "vat_id",

        "northdata_wz_code",
        "openregister_wz_codes",
        "industry_codes",

        "northdata_business_model",
        "openregister_purpose",

        "northdata_revenue_eur",
        "openregister_revenue_eur",

        "northdata_employees",
        "openregister_employees",

        "northdata_balance_sheet_total_eur",
        "openregister_balance_sheet_total_eur",

        "northdata_net_income_eur",
        "openregister_net_income_eur",

        "northdata_equity_eur",
        "openregister_equity_eur",

        "northdata_cash_eur",
        "openregister_cash_eur",

        "northdata_liabilities_eur",
        "openregister_liabilities_eur",

        "northdata_financials_date",
        "openregister_financials_date",

        "number_of_shareholders",
        "natural_person_shareholder_count",
        "legal_person_shareholder_count",
        "youngest_shareholder_age",
        "oldest_shareholder_age",

        "source",
        "company_info_enriched_at",
        "financials_enriched_at",
        "ownership_enriched_at",
        "ubos_enriched_at",
        "created_at",
        "updated_at",
    ],

    "Financials": [
        "openregister_company_id",
        "company_name",

        "northdata_revenue_eur",
        "openregister_revenue_eur",

        "northdata_employees",
        "openregister_employees",

        "northdata_balance_sheet_total_eur",
        "openregister_balance_sheet_total_eur",

        "northdata_net_income_eur",
        "openregister_net_income_eur",

        "northdata_equity_eur",
        "openregister_equity_eur",

        "northdata_cash_eur",
        "openregister_cash_eur",

        "northdata_liabilities_eur",
        "openregister_liabilities_eur",

        "northdata_financials_date",
        "openregister_financials_date",

        "report_count",
        "latest_report_start_date",
        "latest_report_end_date",
        "api_status",
        "notes",
        "enriched_at",
        "updated_at",
    ],

    "Shareholders": [
        "openregister_company_id",
        "company_name",
        "shareholder_name",
        "owner_type",
        "relation_type",
        "percentage_share",
        "nominal_share_eur",
        "age",
        "date_of_birth",
        "natural_person_full_name",
        "natural_person_first_name",
        "natural_person_last_name",
        "legal_person_name",
        "owner_city",
        "owner_country",
        "best_available",
        "api_status",
        "retrieved_at",
        "updated_at",
    ],

    "UBO Control Chain": [
        "openregister_company_id",
        "company_name",
        "ubo_name",
        "ubo_type",
        "percentage_share",
        "max_percentage_share",
        "age",
        "date_of_birth",
        "natural_person_full_name",
        "natural_person_first_name",
        "natural_person_last_name",
        "legal_person_name",
        "ubo_city",
        "ubo_country",
        "api_status",
        "enriched_at",
        "updated_at",
    ],

    "Company Models": [
        "openregister_company_id",
        "company_name",
        "website",
        "model_provider",
        "model_name",
        "business_segment",
        "business_segment_2",
        "business_model",
        "summary",
        "api_status",
        "notes",
        "created_at",
        "updated_at",
    ],

    "Fit Scores": [
        "openregister_company_id",
        "company_name",
        "fit_score",
        "fit_label",
        "fit_comment",
        "succession_signal",
        "financial_signal",
        "shareholder_signal",
        "risk_flags",
        "api_status",
        "created_at",
        "updated_at",
    ],
}


def safe(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def dedupe_preserve_order(values: list[Any]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []

    for value in values or []:
        item = safe(value)

        if not item or item in seen:
            continue

        seen.add(item)
        output.append(item)

    return output


def chunked(values: list[str], size: int):
    for i in range(0, len(values), size):
        yield values[i : i + size]


def fetch_all_rows_paginated(
    supabase,
    table_name: str,
    chunk_size: int = 1000,
    hard_cap: int = 50000,
) -> list[dict[str, Any]]:
    all_rows: list[dict[str, Any]] = []
    start = 0

    while len(all_rows) < hard_cap:
        end = min(start + chunk_size - 1, hard_cap - 1)

        response = supabase.table(table_name).select("*").range(start, end).execute()
        rows = getattr(response, "data", None) or []

        if not rows:
            break

        all_rows.extend(rows)

        if len(rows) < chunk_size:
            break

        start += chunk_size

    return all_rows


def fetch_rows_for_ids(
    supabase,
    table_name: str,
    column_name: str,
    ids: list[str],
    chunk_size: int = 100,
) -> list[dict[str, Any]]:
    ids = dedupe_preserve_order(ids)

    if not ids:
        return []

    collected: list[dict[str, Any]] = []
    seen: set[str] = set()

    for chunk in chunked(ids, chunk_size):
        try:
            res = supabase.table(table_name).select("*").in_(column_name, chunk).execute()
            rows = getattr(res, "data", None) or []
        except Exception:
            rows = []

        for row in rows:
            key = (
                safe(row.get("id"))
                or safe(row.get("openregister_company_id"))
                + safe(row.get("owner_key"))
                + safe(row.get("ubo_key"))
                + safe(row.get("model_provider"))
            )

            if key in seen:
                continue

            seen.add(key)
            collected.append(row)

    return collected


def _safe_cell(value: Any, column_name: str | None = None) -> Any:
    if column_name in {"industry_codes", "openregister_wz_codes"}:
        value = format_industry_codes(value)
    else:
        value = flatten_for_sheet(value)

    if column_name in {
        "main_ubo_max_percentage_share",
        "max_percentage_share",
        "ubo_max_percentage_share",
    } and value not in (None, ""):
        try:
            value = round(float(value), 2)
        except Exception:
            pass

    if isinstance(value, str) and len(value) > 32000:
        return value[:32000] + "… [truncated; full value remains in Supabase]"

    return value


def rows_to_values(
    rows: list[dict[str, Any]],
    *,
    preferred_columns: list[str] | None = None,
    exclude_columns: set[str] | None = None,
) -> list[list[Any]]:
    exclude = set(EXPORT_EXCLUDED_COLUMNS) | set(exclude_columns or set())
    rows = rows or []

    cleaned = [
        {key: value for key, value in row.items() if key not in exclude}
        for row in rows
    ]

    if not cleaned:
        return [["No rows"]]

    if preferred_columns:
        columns = [column for column in preferred_columns if any(column in row for row in cleaned)]

        extra: list[str] = []

        for row in cleaned:
            for key in row.keys():
                if key not in columns and key not in extra:
                    extra.append(key)

        columns.extend(extra)

    else:
        columns = []

        for row in cleaned:
            for key in row.keys():
                if key not in columns:
                    columns.append(key)

    values = [[nice_sheet_header(column) for column in columns]]

    for row in cleaned:
        values.append([_safe_cell(row.get(column), column) for column in columns])

    return values


def style_sheet(ws):
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = TITLE_FILL
            cell.font = WHITE_BOLD
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER

    ws.freeze_panes = "A2"

    try:
        ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass

    auto_fit_columns(ws)


def auto_fit_columns(ws, max_width: int = 45):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = 0

        for cell in col_cells:
            if cell.value is None:
                continue

            max_len = max(max_len, min(len(str(cell.value)), max_width))

        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def write_sheet(wb: Workbook, title: str, values: list[list[Any]]) -> None:
    ws = wb.create_sheet(title=title[:31])

    for row in values:
        ws.append(row)

    style_sheet(ws)


def _select_financial_rows_for_ids(
    financial_rows: list[dict[str, Any]],
    register_ids: list[str],
) -> list[dict[str, Any]]:
    wanted = set(register_ids)

    return [
        row
        for row in financial_rows
        if safe(row.get("company_register_id")) in wanted
        or safe(row.get("register_id")) in wanted
    ]


def build_filtered_workbook_bytes(
    supabase,
    register_ids: list[str],
    overview_rows: list[dict[str, Any]] | None = None,
    log_callback=None,
) -> dict[str, Any]:
    register_ids = dedupe_preserve_order(register_ids)

    if not register_ids:
        raise ValueError("No companies selected for export.")

    if log_callback:
        log_callback(f"Fetching workbook data for {len(register_ids)} companies...")

    if overview_rows is None:
        overview_rows = fetch_rows_for_ids(
            supabase,
            "master_overview",
            "register_id",
            register_ids,
        )

    companies = fetch_rows_for_ids(
        supabase,
        "companies",
        "register_id",
        register_ids,
    )

    financials_all = _fetch_financials_sheet_rows(supabase)
    financials = _select_financial_rows_for_ids(financials_all, register_ids)

    shareholders = fetch_rows_for_ids(
        supabase,
        "shareholders",
        "company_register_id",
        register_ids,
    )

    ubos = fetch_rows_for_ids(
        supabase,
        "company_ubos",
        "company_register_id",
        register_ids,
    )

    models = fetch_rows_for_ids(
        supabase,
        "company_models",
        "company_register_id",
        register_ids,
    )

    scores = fetch_rows_for_ids(
        supabase,
        "company_fit_scores",
        "company_register_id",
        register_ids,
    )

    logs = fetch_rows_for_ids(
        supabase,
        "processing_logs",
        "company_register_id",
        register_ids,
    )

    wb = Workbook()
    wb.remove(wb.active)

    sheet_specs = [
        ("Overview", overview_rows),
        ("Companies", companies),
        ("Financials", financials),
        ("Shareholders", shareholders),
        ("UBO Control Chain", ubos),
        ("Company Models", models),
        ("Fit Scores", scores),
        ("Processing Logs", logs),
    ]

    for title, rows in sheet_specs:
        write_sheet(
            wb,
            title,
            rows_to_values(
                rows,
                preferred_columns=PREFERRED_COLUMN_ORDER.get(title),
            ),
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    table_counts = {title: len(rows or []) for title, rows in sheet_specs}

    return {
        "workbook_bytes": buffer.getvalue(),
        "table_counts": table_counts,
        "selected_register_ids": register_ids,
        "company_rows": len(companies),
    }


def apply_numeric_filter(
    df: pd.DataFrame,
    column: str,
    operator: str,
    value1: float | None,
    value2: float | None = None,
) -> pd.DataFrame:
    if column not in df.columns or operator == "Ignore" or value1 is None:
        return df

    series = pd.to_numeric(df[column], errors="coerce")

    if operator == "=":
        return df[series == value1]

    if operator == ">":
        return df[series > value1]

    if operator == ">=":
        return df[series >= value1]

    if operator == "<":
        return df[series < value1]

    if operator == "<=":
        return df[series <= value1]

    if operator == "Between":
        return df[(series >= value1) & (series <= (value2 if value2 is not None else value1))]

    return df
